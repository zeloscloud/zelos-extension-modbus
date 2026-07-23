#!/usr/bin/env python3
"""Convert a Modbus register-map spreadsheet into the extension's JSON format.

Expected spreadsheet columns (first row is header):

    Register Index | Register Name | Read/Write | Resettable via Modbus
        | Value | Units | Value Description

Each data row becomes its own single-register uint16 holding event. The event
name and its lone field name are both the register name, snake_cased and made
unique when a name repeats (`exercise_time`, `exercise_time_2`, ...), so the
signal tree reads cleanly.

Unit suffixes ending in `x10` (e.g. `v dc x10`) are interpreted as scale=0.1
with the `x10` stripped from the displayed unit.

Usage:
    uv run --with openpyxl python scripts/xlsx_to_register_map.py \\
        ~/Downloads/RCM_GMS_Register_Map_NoGrouping.xlsx \\
        examples/rcm_gms_register_map.json
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    sys.stderr.write(
        "openpyxl not installed. Run with: "
        "`uv run --with openpyxl python scripts/xlsx_to_register_map.py ...`\n"
    )
    raise


_SNAKE_RE = re.compile(r"[^a-z0-9]+")


def to_snake(name: str) -> str:
    """Convert an arbitrary label to snake_case."""
    s = _SNAKE_RE.sub("_", name.strip().lower()).strip("_")
    return s or "unnamed"


def parse_unit(unit: str | None) -> tuple[str, float]:
    """Return (clean_unit, scale) from a spreadsheet unit cell.

    `v dc x10` style units indicate the raw register is the value * 10
    (so the decoded/displayed value needs scale 0.1).
    """
    if not unit:
        return "", 1.0
    u = str(unit).strip()
    m = re.match(r"^(.*?)\s*[x*]\s*(\d+(?:\.\d+)?)\s*$", u, flags=re.IGNORECASE)
    if m:
        base = m.group(1).strip()
        factor = float(m.group(2))
        if factor != 0:
            return base, 1.0 / factor
    # "0.05 Hours/bit" style: raw counts of a fixed quantum
    m = re.match(r"^(\d+(?:\.\d+)?)\s*(.*?)\s*/\s*bit\s*$", u, flags=re.IGNORECASE)
    if m:
        return m.group(2).strip(), float(m.group(1))
    return u, 1.0


def parse_row(row: tuple[Any, ...]) -> dict[str, Any] | None:
    """Convert one spreadsheet row into a register dict (or None to skip)."""
    if not row or row[0] is None:
        return None
    idx, name, rw, _resettable, value, unit, desc = (row + (None,) * 7)[:7]
    # bool is an int subclass, so guard it before the int check: a TRUE/FALSE
    # cell is not a register index.
    if isinstance(idx, bool):
        return None
    # Excel sheets often store some indices as text cells; accept "8505" but
    # still skip range rows like "10010-10015" (reserved/log blocks).
    if isinstance(idx, str) and idx.strip().isdigit():
        idx = int(idx.strip())
    # Whole-number float cells (8505.0) are valid indices; fractional ones are not.
    if isinstance(idx, float) and idx.is_integer():
        idx = int(idx)
    if not isinstance(idx, int):
        return None
    if not name:
        return None

    clean_unit, scale = parse_unit(unit)

    rw_str = str(rw).strip().upper() if rw else ""
    writable = rw_str in ("W", "R/W", "RW")

    # Combine Value (e.g. "0-255:0-255") and Value Description into one string
    desc_parts: list[str] = []
    if value not in (None, ""):
        desc_parts.append(f"[{value}]")
    if desc:
        desc_parts.append(str(desc).strip())
    description = " ".join(desc_parts)

    reg: dict[str, Any] = {
        "name": str(name).strip(),  # replaced with the unique event name later
        "address": int(idx),
        "type": "holding",
        "datatype": "uint16",
        "writable": writable,
    }
    if clean_unit:
        reg["unit"] = clean_unit
    if scale != 1.0:
        reg["scale"] = scale
    if description:
        reg["description"] = description
    return reg


def build_register_map(
    rows: list[tuple[Any, ...]],
    device_name: str,
    description: str,
) -> dict[str, Any]:
    """Turn parsed rows into the extension's register-map JSON structure.

    Each row becomes its own single-register event; the event name and its lone
    field name are the snake_cased register name, uniquified (`_2`, `_3`, ...)
    when a name repeats. Identical duplicate rows (same name/address/type) are
    dropped.
    """
    seen: set[tuple[str, int, str]] = set()
    used_events: set[str] = set()
    events: dict[str, list[dict[str, Any]]] = {}

    for row in rows:
        reg = parse_row(row)
        if reg is None:
            continue
        # Some sheets repeat identical rows (e.g. a summary block); keep the first.
        key = (reg["name"], reg["address"], reg["type"])
        if key in seen:
            continue
        seen.add(key)

        event = to_snake(reg["name"])
        base = event
        n = 2
        while event in used_events:
            event = f"{base}_{n}"
            n += 1
        used_events.add(event)

        reg["name"] = event
        events[event] = [reg]

    return {
        "name": device_name,
        "description": description,
        "events": events,
    }


def load_rows(ws: Any) -> list[tuple[Any, ...]]:
    """Read the data rows (everything after the header) from an open worksheet."""
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return []
    header = tuple((c or "").strip() if isinstance(c, str) else c for c in all_rows[0])
    expected = ("Register Index", "Register Name", "Read/Write")
    if header[:3] != expected:
        sys.stderr.write(
            f"warning: first 3 header columns are {header[:3]!r}, expected {expected!r}\n"
        )
    return all_rows[1:]


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("input", type=Path, help="Path to .xlsx register map")
    ap.add_argument("output", type=Path, help="Path to write JSON register map")
    ap.add_argument("--sheet", default=None, help="Worksheet name (default: first sheet)")
    ap.add_argument("--name", default=None, help="Device name (default: sheet name lowercased)")
    ap.add_argument("--description", default=None, help="Register-map description")
    args = ap.parse_args(argv)

    if not args.input.exists():
        sys.stderr.write(f"input not found: {args.input}\n")
        return 2

    # Load once and resolve the worksheet up front so the sheet title can be
    # threaded through without a second workbook load.
    wb = openpyxl.load_workbook(args.input, data_only=True)
    if args.sheet is not None and args.sheet not in wb.sheetnames:
        sys.stderr.write(f"sheet {args.sheet!r} not found. Available sheets: {wb.sheetnames}\n")
        return 2
    sheet_name = args.sheet or wb.sheetnames[0]

    rows = load_rows(wb[sheet_name])
    device_name = args.name or to_snake(sheet_name)
    description = args.description or f"Generated from {args.input.name}"

    register_map = build_register_map(rows, device_name=device_name, description=description)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w") as f:
        json.dump(register_map, f, indent=2)
        f.write("\n")

    total = sum(len(regs) for regs in register_map["events"].values())
    print(
        f"wrote {args.output} — {len(register_map['events'])} events, {total} registers",
        file=sys.stderr,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
